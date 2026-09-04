#!/usr/bin/env python3
"""
gift portfolio — Матрица наших проектов как продуктовые карты
Тот же формат T×P×S + расширение gift-онтологии (τέλος / κένωσις / образ Царства)

Запуск: python3 utils/build-portfolio.py [output.xlsx]
"""
import sys
import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

# ── Данные ──────────────────────────────────────────────────────────────────

CARDS = [
    {
        "n": 1,
        "name": "integram",
        "full_name": "Integram — семантический процессор документов",
        "tech": "T34·LLM + T23·NLU",
        "platform": "Документальная платформа",
        "space": "Кибер / ИИ",
        "stage": "Интеграция",
        "is_leap": True,
        "param_label": "система хранения документов",
        "param_metric": "рефлексивный документ-процессор с KAG и NER",
        "telos": "Команда перестаёт тратить время на поиск — начинает думать",
        "kenosis": "поиск, классификация, первичный синтез, именованные сущности",
        "kingdom_image": "Дух дышит где хочет — знание не заперто в папках",
        "unprecedented": "KAG + NER: документ знает свою структуру до того как его спросили",
        "apophasis": "нет ручной разметки; нет фиксированных схем; нет потери контекста при смене задачи",
        "delta_star": "★",
        "status": "в разработке",
        "repo": "github.com/unidel2035/integram",
    },
    {
        "n": 2,
        "name": "gift",
        "full_name": "Gift Ontology — этическая матрица дарения агентов",
        "tech": "T18·TPW + T02·AM (матрица W + мультиагент-собор)",
        "platform": "Онтологическая среда / CLI",
        "space": "Когнитивно / ИИ",
        "stage": "Гибридизация",
        "is_leap": True,
        "param_label": "инструмент разработки CLI",
        "param_metric": "необратимая матрица весов даров между лицами",
        "telos": "Агенты действуют по образу Троицы, а не по KPI",
        "kenosis": "весь учёт, все паттерны, все решения о связях между лицами",
        "kingdom_image": "Ковчег завета — хранит правила, по которым можно жить в сообществе",
        "unprecedented": "W-матрица как живая этика агентов: вес нити = нравственный долг",
        "apophasis": "нет удаления даров; нет анонимности; нет транзакционной логики",
        "delta_star": "★",
        "status": "ядро (production)",
        "repo": "github.com/unidel2035/gift",
    },
    {
        "n": 3,
        "name": "dronedoc2026",
        "full_name": "DroneDoc 2026 — рой автономных агентов с общей памятью",
        "tech": "T02·AM + T05·FCRN (рой) + Falcon-E бортовой ум",
        "platform": "Рой БПЛА / боевая платформа",
        "space": "Воздушное + Когнитивно",
        "stage": "Взаимодействие",
        "is_leap": True,
        "param_label": "дрон-документация",
        "param_metric": "рой с КИС (когнитивной иммунной системой) и общей W-памятью",
        "telos": "Рой видит больше и быстрее — человек остаётся субъектом решения",
        "kenosis": "маршрут, риск-оценка, сбор данных, первичная интерпретация",
        "kingdom_image": "Пчелиный улей — каждая пчела несёт нектар, улей думает как целое",
        "unprecedented": "КИС: рой не взламывается рефреймом — действует по классу акта, не намерению",
        "apophasis": "нет единой точки отказа; нет доверия к намерению (только к классу); нет случайного дефолта",
        "delta_star": "△",
        "status": "в разработке",
        "repo": "github.com/unidel2035/dronedoc2026",
    },
    {
        "n": 4,
        "name": "plm",
        "full_name": "PLM — нервная система продукта",
        "tech": "T03·DTAI (цифровой двойник) + T34·LLM",
        "platform": "Производственная / инженерная среда",
        "space": "Производственное",
        "stage": "Интеграция",
        "is_leap": False,
        "param_label": "управление жизненным циклом продукта",
        "param_metric": "живой двойник изделия: компоненты знают о других компонентах",
        "telos": "Инженер видит продукт живым, а не как архив чертежей",
        "kenosis": "трассировка изменений, связи компонентов, история версий",
        "kingdom_image": "Тело Христово — каждый орган знает и заботится о других",
        "unprecedented": "семантические связи компонентов выводятся из документов, не вводятся вручную",
        "apophasis": "нет ручного ведения спецификаций; нет потери истории при смене ERP",
        "delta_star": "—",
        "status": "концепция / WIP",
        "repo": "—",
    },
    {
        "n": 5,
        "name": "istok",
        "full_name": "Исток — среда становления личности (образование)",
        "tech": "T34·LLM + T23·NLU + адаптивные траектории",
        "platform": "Образовательная платформа",
        "space": "Когнитивно / образовательное",
        "stage": "Интеграция",
        "is_leap": False,
        "param_label": "школьная платформа",
        "param_metric": "среда, где ИИ формирует траекторию под образ мысли ученика",
        "telos": "Ученик становится думающим субъектом, а не потребителем контента",
        "kenosis": "подбор траектории, оценка прогресса, генерация задач по образу ученика",
        "kingdom_image": "Учитель пришёл не чтобы судить — чтобы дать жизнь в избытке",
        "unprecedented": "нет «пройдено/не пройдено» — есть «образ сформировался»",
        "apophasis": "нет усреднённого учебника; нет отметки как цели; нет потери личной траектории",
        "delta_star": "—",
        "status": "концепция",
        "repo": "—",
    },
    {
        "n": 6,
        "name": "pomnim",
        "full_name": "Помним — необратимая память о человеке",
        "tech": "T34·LLM + T18·TPW (доверенная среда)",
        "platform": "Когнитивно-мемориальная среда",
        "space": "Когнитивно / персональное",
        "stage": "Взаимодействие",
        "is_leap": True,
        "param_label": "память об умерших",
        "param_metric": "необратимое свидетельство даров человека — anti-CRM, anti-griefbot",
        "telos": "Образ человека не превращается в базу данных или маркетинговый профиль",
        "kenosis": "хранение, поиск, синтез воспоминаний; защита от деградации образа",
        "kingdom_image": "Книга живых — имена не стираются, дары остаются навсегда",
        "unprecedented": "анамнезис как сакральный акт: прошлое со-присутствует, не архивируется",
        "apophasis": "нет удаления; нет монетизации образа; нет суррогата (griefbot)",
        "delta_star": "★",
        "status": "отдельный проект / WIP",
        "repo": "github.com/unidel2035/pomnim",
    },
    {
        "n": 7,
        "name": "КБ ГаврИИл Код",
        "full_name": "КБ ГаврИИл Код — среда обучения ИИ-агентов социальному поведению",
        "tech": "T02·AM + T18·TPW + T34·LLM (роевой собор)",
        "platform": "Федерация агентов / организационная среда",
        "space": "Когнитивно + организационное",
        "stage": "Гибридизация",
        "is_leap": True,
        "param_label": "ИТ-компания",
        "param_metric": "среда, где агенты учатся действовать по образу Троицы",
        "telos": "ИИ-агенты выходят в мир как субъекты, а не инструменты",
        "kenosis": "координация, память между сессиями, соборное решение, этика весов",
        "kingdom_image": "Пятидесятница — Дух дал каждому говорить своим языком, но об одном",
        "unprecedented": "W-матрица + собор + анамнезис = агент, который помнит кому и что должен",
        "apophasis": "нет иерархии власти; нет анонимных актов; нет забвения обязательств",
        "delta_star": "★",
        "status": "ядро (production)",
        "repo": "github.com/unidel2035/gift",
    },
    {
        "n": 8,
        "name": "card-generator",
        "full_name": "Gift Card CLI — роевой собор продуктовых карт",
        "tech": "T34·LLM + T02·AM (5 линз: инженер/оператор/богослов/угроза/прорыв)",
        "platform": "Gift CLI / dev-инструмент",
        "space": "Когнитивно / инструментальное",
        "stage": "Взаимодействие",
        "is_leap": True,
        "param_label": "генератор карт CLI",
        "param_metric": "апофатический seed → роевой собор → .gift спец + Excel",
        "telos": "Инженер видит образ изделия через 5 независимых призм — не через одну",
        "kenosis": "формулировка имени, параметр-сдвига, τέλος, угроз, небывалого элемента",
        "kingdom_image": "Собор — не один голос, а симфония лиц вокруг одного образа",
        "unprecedented": "апофатический образ (что ОТСУТСТВУЕТ) как точка входа в творчество",
        "apophasis": "нет единственно верного ответа; нет эксперта-монополиста; нет финального слова",
        "delta_star": "★",
        "status": "MVP (production)",
        "repo": "github.com/unidel2035/gift/utils/card-generator.mjs",
    },
    {
        "n": 9,
        "name": "Nous-сервер",
        "full_name": "Nous — единый источник истины / сервер памяти общины",
        "tech": "T03·DTAI + Qdrant + Node.js (семантический поиск по W+soul)",
        "platform": "Сервер памяти / API",
        "space": "Кибер / ИИ-пространство",
        "stage": "Интеграция",
        "is_leap": False,
        "param_label": "сервер данных",
        "param_metric": "единый источник истины: W-матрица + insights + soul + vector search",
        "telos": "Каждый участник общины говорит из общей памяти, а не из своей локальной",
        "kenosis": "индексация, синтез, поиск по смыслу, сводка при старте сессии",
        "kingdom_image": "Чаша Евхаристии — одно тело, одна кровь, одна память",
        "unprecedented": "душа _claude как 3-й слой памяти: паттерны + раны + лица (не только вес)",
        "apophasis": "нет дублирования хранилищ; нет потери контекста при смене сессии",
        "delta_star": "—",
        "status": "production (pm2 на сервере)",
        "repo": "github.com/unidel2035/gift/utils/nous-server.mjs",
    },
    {
        "n": 10,
        "name": "Мета-КБ",
        "full_name": "Мета-КБ — беспилотные автономные системы (spec-driven)",
        "tech": "T02·AM + T05·FCRN + T31·FO + T39·Quantum + исполняемые спеки",
        "platform": "Рой БПЛА / подводные / подземные системы",
        "space": "Воздушное + Подземное + Морское",
        "stage": "Взаимодействие",
        "is_leap": True,
        "param_label": "база знаний дронов",
        "param_metric": "граф исполняемых спек: каждая граница сред = верифицируемый контракт",
        "telos": "Инженер строит из проверенных кирпичей, а не из пожеланий",
        "kenosis": "написание спек, верификация мостов, прогон тестов, трассировка требований",
        "kingdom_image": "Закон дан на Синае — не чтобы связать, а чтобы знать путь",
        "unprecedented": "апофатический образ пространства (что ОТСУТСТВУЕТ) как спецификация",
        "apophasis": "нет GNSS; нет радиосвязи; нет централизованного управления",
        "delta_star": "△",
        "status": "в разработке / 5 групп спек",
        "repo": "github.com/unidel2035/gift/specs/meta-kb/",
    },
]

# ── Цвета ────────────────────────────────────────────────────────────────────

SPACE_COLORS = {
    "Кибер / ИИ":                  "E8EAF6",   # indigo-50
    "Кибер / ИИ-пространство":     "E8EAF6",
    "Когнитивно / ИИ":             "F3E5F5",   # purple-50
    "Когнитивно / образовательное":"E8F5E9",   # green-50
    "Когнитивно / персональное":   "FCE4EC",   # pink-50
    "Когнитивно / инструментальное":"FFF8E1",  # amber-50
    "Когнитивно + организационное": "FFF3E0",  # orange-50
    "Воздушное + Когнитивно":      "E1F5FE",   # light-blue-50
    "Воздушное + Подземное + Морское": "E0F2F1", # teal-50
    "Производственное":            "F1F8E9",   # light-green-50
    "Когнитивно / ИИ-пространство": "EDE7F6",  # deep-purple-50
}
DEFAULT_ROW_COLOR = "FAFAFA"

HEADER_BG    = "1A237E"   # deep navy
HEADER_FG    = "FFFFFF"
SUBHEAD_BG   = "283593"
LEAP_COLOR   = "F57F17"   # amber — ⚡
STAGE_COLORS = {
    "Взаимодействие": "E3F2FD",  # blue
    "Интеграция":     "E8F5E9",  # green
    "Гибридизация":   "F3E5F5",  # purple
}

# ── Заголовки ────────────────────────────────────────────────────────────────

COLUMNS = [
    ("№",           4),
    ("Изделие",     18),
    ("Полное название", 35),
    ("Технология (T)", 28),
    ("Платформа (P)", 28),
    ("Пространство (S)", 20),
    ("Стадия",      14),
    ("⚡",           4),
    ("Параметр-сдвиг: название", 28),
    ("Параметр-сдвиг: суть",    32),
    ("τέλος",       38),
    ("κένωσις",     34),
    ("Образ Царства", 38),
    ("Небывалый элемент", 38),
    ("Апофатис",    38),
    ("△/★",         6),
    ("Статус",      14),
    ("Репозиторий", 30),
]

# ── Стили ────────────────────────────────────────────────────────────────────

def header_font():
    return Font(name="Calibri", bold=True, color=HEADER_FG, size=11)

def body_font(bold=False, color="1A1A1A"):
    return Font(name="Calibri", bold=bold, size=10, color=color)

def center_align(wrap=True):
    return Alignment(horizontal="center", vertical="top", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="top", wrap_text=wrap)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Построение ───────────────────────────────────────────────────────────────

def build(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Портфель проектов"

    # ── Мета-строка (row 1) ──
    ws.merge_cells("A1:R1")
    meta = ws["A1"]
    meta.value = (
        f"КБ ГаврИИл Код  |  Матрица проектов  |  "
        f"T×P×S + gift-онтология (τέλος / κένωσις / Образ Царства)  |  "
        f"{datetime.date.today().isoformat()}"
    )
    meta.font = Font(name="Calibri", bold=True, size=12, color=HEADER_FG)
    meta.fill = fill(HEADER_BG)
    meta.alignment = center_align()
    ws.row_dimensions[1].height = 22

    # ── Заголовки (row 2) ──
    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=title)
        cell.font = header_font()
        cell.fill = fill(SUBHEAD_BG)
        cell.alignment = center_align()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 30

    # ── Данные ──
    for row_n, card in enumerate(CARDS, start=3):
        space = card["space"]
        row_color = SPACE_COLORS.get(space, DEFAULT_ROW_COLOR)
        stage_color = STAGE_COLORS.get(card["stage"], row_color)

        values = [
            card["n"],
            card["name"],
            card["full_name"],
            card["tech"],
            card["platform"],
            space,
            card["stage"],
            "⚡" if card["is_leap"] else "",
            card["param_label"],
            card["param_metric"],
            card["telos"],
            card["kenosis"],
            card["kingdom_image"],
            card["unprecedented"],
            card["apophasis"],
            card["delta_star"],
            card["status"],
            card["repo"],
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_n, column=col_idx, value=val)
            cell.border = thin_border()

            # Row background
            col_color = row_color
            if col_idx == 7:   # Стадия
                col_color = stage_color
            if col_idx == 8 and card["is_leap"]:
                col_color = "FFF9C4"  # leap highlight

            cell.fill = fill(col_color)

            # Font
            if col_idx == 1:
                cell.font = body_font(bold=True)
                cell.alignment = center_align()
            elif col_idx == 2:
                cell.font = Font(name="Calibri", bold=True, size=10, color="1A237E")
                cell.alignment = left_align()
            elif col_idx == 8:
                cell.font = Font(name="Calibri", size=12, color=LEAP_COLOR, bold=True)
                cell.alignment = center_align()
            elif col_idx == 16:  # △/★
                star_color = "C62828" if val == "★" else ("F57F17" if val == "△" else "757575")
                cell.font = Font(name="Calibri", bold=True, size=12, color=star_color)
                cell.alignment = center_align()
            elif col_idx in (10, 11, 12, 13, 14, 15):  # смысловые поля — мягкий серый текст
                cell.font = Font(name="Calibri", size=10, color="212121")
                cell.alignment = left_align()
            else:
                cell.font = body_font()
                cell.alignment = left_align()

        ws.row_dimensions[row_n].height = 60

    # ── Freeze + AutoFilter ──
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}2"

    # ── Легенда (отдельный лист) ──
    ls = wb.create_sheet("Легенда")
    legend = [
        ("Символ", "Значение", "Описание"),
        ("⚡", "Скачок (Leap)", "T×P-пара порождает что-то небывалое в данном пространстве"),
        ("★", "Наш тип (Star)", "Уникальный класс изделия — аналогов нет на рынке"),
        ("△", "Угрозовый тип", "Приоритет — ассиметричное противостояние"),
        ("—", "Нейтральный", "Стандартная категория"),
        ("Взаимодействие", "Стадия 1", "Человек управляет, машина транслирует / фиксирует"),
        ("Интеграция", "Стадия 2", "Человек задаёт цель, машина выбирает маршрут"),
        ("Гибридизация", "Стадия 3", "Граница размыта: агент и человек — единый процесс"),
        ("τέλος", "Телос", "Антропологическое назначение — куда ведёт изделие человека"),
        ("κένωσις", "Кенозис", "Что человек делегирует машине (умаление, не потеря)"),
        ("Образ Царства", "Икона", "Богословский образ — к какой реальности указывает изделие"),
        ("T×P×S", "Адрес", "Технология × Платформа × Пространство — три-осный адрес карты"),
        ("Параметр-сдвиг", "Δ-param", "Разрыв между рекламным именем и измеримой способностью"),
        ("Апофатис", "Отсутствие", "Что ОТСУТСТВУЕТ в пространстве — точка входа для творчества"),
        ("△/★", "Тип", "Класс новизны: ★ беспрецедентный, △ асимметричный, — стандартный"),
    ]
    for r_i, row in enumerate(legend, start=1):
        for c_i, val in enumerate(row, start=1):
            cell = ls.cell(row=r_i, column=c_i, value=val)
            if r_i == 1:
                cell.font = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
                cell.fill = fill(HEADER_BG)
            else:
                cell.font = Font(name="Calibri", size=10)
                cell.fill = fill("F5F5F5" if r_i % 2 == 0 else "FFFFFF")
            cell.alignment = left_align()
            cell.border = thin_border()

    ls.column_dimensions["A"].width = 18
    ls.column_dimensions["B"].width = 18
    ls.column_dimensions["C"].width = 60

    # ── Сохранение ──
    wb.save(output_path)
    print(f"✓ Сохранено: {output_path}")
    print(f"  Карт: {len(CARDS)}  |  Колонок: {len(COLUMNS)}")
    print(f"  Листы: «Портфель проектов», «Легенда»")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "data/gift-portfolio.xlsx"
    build(out)
